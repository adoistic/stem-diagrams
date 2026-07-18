"""Build and upload the shareable gallery.

One page at v3/gallery/index.html on the R2 public URL: a live status header
(elapsed since start, progress, rolling ETA, all in the viewer's timezone with
a dropdown to change it) plus a category-filtered grid of every diagram.
Reads v3/gallery/data.json, which the running pipeline regenerates as it works;
the page also re-fetches it every 60s so the numbers stay live.
"""

import json
import time
from pathlib import Path

import r2

R2_PUBLIC = "https://pub-4b177d3c9b154dfeb08296540e8242ee.r2.dev"
DISPLAY = {
    "semiconductor_engineering": "Semiconductor Engineering",
    "manufacturing_engineering": "Manufacturing Engineering",
    "robotics_automation": "Robotics & Automation",
    "utilities_power_systems": "Utilities & Power Systems",
    "telecommunications": "Telecommunications",
}
WORK = Path(__file__).resolve().parent / "work"

INDEX_HTML = r"""<!DOCTYPE html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>STEM Diagrams - live</title>
<style>
:root{--bg:#0f1420;--card:#161d2b;--ink:#e8ecf1;--muted:#9aa4b2;--line:#2a3140;--acc:#2a9d8f}
@media(prefers-color-scheme:light){:root{--bg:#fff;--card:#f7f9fb;--ink:#1a1a2e;--muted:#5a6472;--line:#e6e9ee}}
*{box-sizing:border-box}body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:var(--bg);color:var(--ink)}
header{padding:22px 20px 6px;max-width:1200px;margin:0 auto}
h1{margin:0;font-size:1.5rem}
.stats{max-width:1200px;margin:10px auto 0;padding:0 20px;display:grid;
grid-template-columns:repeat(4,1fr);gap:12px}
.stat{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:12px 14px}
.stat .l{color:var(--muted);font-size:.72rem;text-transform:uppercase;letter-spacing:.04em}
.stat .v{font-size:1.15rem;font-weight:700;margin-top:3px}
.stat .v.acc{color:var(--acc)}
.pbar{max-width:1200px;margin:12px auto 0;padding:0 20px}
.pbar .track{height:8px;background:var(--card);border:1px solid var(--line);border-radius:6px;overflow:hidden}
.pbar .fill{height:100%;background:var(--acc);width:0;transition:width .5s}
.tzrow{max-width:1200px;margin:10px auto 0;padding:0 20px;color:var(--muted);font-size:.82rem;display:flex;gap:8px;align-items:center;flex-wrap:wrap}
select{background:var(--card);color:var(--ink);border:1px solid var(--line);border-radius:8px;padding:5px 8px;font-size:.82rem}
a.exp{margin-left:auto;background:var(--acc);color:#fff;border:none;border-radius:8px;padding:7px 14px;font-weight:700;font-size:.82rem;text-decoration:none;cursor:pointer}
a.exp:hover{opacity:.9}
.bar{position:sticky;top:0;background:var(--bg);border-bottom:1px solid var(--line);z-index:5;padding:12px 20px;margin-top:12px}
.bar .wrap{max-width:1200px;margin:0 auto;display:flex;flex-wrap:wrap;gap:8px;align-items:center}
button.cat{border:1px solid var(--line);background:var(--card);color:var(--muted);padding:7px 14px;border-radius:20px;font-weight:600;font-size:.85rem;cursor:pointer}
button.cat.on{background:var(--acc);color:#fff;border-color:var(--acc)}
main{max-width:1200px;margin:0 auto;padding:16px 20px 60px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(210px,1fr));gap:14px}
.card{background:var(--card);border:1px solid var(--line);border-radius:12px;overflow:hidden;cursor:pointer;transition:transform .12s}.card:hover{transform:translateY(-3px)}
.card img{width:100%;height:150px;object-fit:contain;background:#fff;display:block;padding:6px}
.card .m{padding:8px 10px;font-size:.72rem;color:var(--muted)}
.lb{position:fixed;inset:0;background:rgba(8,12,20,.9);display:none;align-items:center;justify-content:center;z-index:20;padding:20px}.lb.on{display:flex}
.lb .inner{max-width:900px;width:100%;max-height:92vh;overflow:auto;background:var(--bg);border-radius:14px}
.lb img{width:100%;background:#fff;border-radius:14px 14px 0 0;max-height:70vh;object-fit:contain;padding:10px}
.lb .c{padding:14px 18px}.lb a{color:var(--acc)}
.lb .x{position:fixed;top:16px;right:22px;color:#fff;font-size:2rem;cursor:pointer}
.more{margin:24px auto;display:block;padding:10px 20px;border-radius:10px;border:1px solid var(--line);background:var(--card);color:var(--ink);font-weight:600;cursor:pointer}
@media(max-width:720px){.stats{grid-template-columns:repeat(2,1fr)}}
</style></head><body>
<header><h1>STEM Diagrams</h1></header>
<div class=stats>
 <div class=stat><div class=l>Started</div><div class=v id=s-start>-</div></div>
 <div class=stat><div class=l>Running for</div><div class=v id=s-elapsed>-</div></div>
 <div class=stat><div class=l>Diagrams</div><div class="v acc" id=s-count>-</div></div>
 <div class=stat><div class=l>Est. finish</div><div class=v id=s-eta>-</div></div>
</div>
<div class=pbar><div class=track><div class=fill id=s-fill></div></div></div>
<div class=tzrow><span>Times shown in</span>
 <select id=tz></select><span id=s-sub></span>
 <a class=exp id=exp href="../export/stem_diagrams.xlsx" download>⤓ Export (Excel)</a></div>
<div class=bar><div class=wrap id=cats></div></div>
<main><div class=grid id=grid></div><button class=more id=more style=display:none>Show more</button></main>
<div class=lb id=lb><span class=x onclick="lb.classList.remove('on')">&times;</span>
<div class=inner><img id=lbimg><div class=c id=lbc></div></div></div>
<script>
let DATA=null,cat=null,shown=0,PAGE=60,TZ=null;
const $=id=>document.getElementById(id);
const cats=$('cats'),grid=$('grid'),more=$('more'),lb=$('lb');
function fmtTime(epoch){try{return new Intl.DateTimeFormat([],{dateStyle:'medium',timeStyle:'short',timeZone:TZ}).format(new Date(epoch*1000))}catch(e){return new Date(epoch*1000).toLocaleString()}}
function dur(sec){sec=Math.max(0,Math.floor(sec));const h=Math.floor(sec/3600),m=Math.floor(sec%3600/60),s=sec%60;return (h?h+'h ':'')+(h||m?m+'m ':'')+s+'s'}
function durShort(sec){sec=Math.max(0,Math.floor(sec));const h=Math.floor(sec/3600),m=Math.floor(sec%3600/60);return h?`${h}h ${m}m`:`${m}m`}
function tick(){ if(!DATA)return;
 $('s-elapsed').textContent=dur(Date.now()/1000-DATA.start);
}
function renderStats(){ if(!DATA)return;
 $('s-start').textContent=fmtTime(DATA.start);
 const el=DATA.updated_epoch-DATA.start, rate=el>0?DATA.count/el:0;
 const unlimited=!DATA.target||DATA.target<=0;
 if(unlimited){
  $('s-count').textContent=DATA.count.toLocaleString();
  $('s-fill').style.width='100%';document.getElementById('s-fill').style.opacity='.35';
  $('s-eta').innerHTML=`running<br><span style="font-size:.72rem;color:var(--muted);font-weight:400">${(rate*3600).toFixed(0)}/hr · no fixed target</span>`;
 }else{
  $('s-count').textContent=`${DATA.count.toLocaleString()} / ${DATA.target.toLocaleString()}`;
  $('s-fill').style.width=Math.min(100,DATA.count/DATA.target*100)+'%';
  if(DATA.count>=DATA.target){$('s-eta').textContent='Complete';}
  else if(rate>0){const rem=(DATA.target-DATA.count)/rate;
   $('s-eta').innerHTML=`${fmtTime(DATA.updated_epoch+rem)}<br><span style="font-size:.72rem;color:var(--muted);font-weight:400">~${durShort(rem)} left · ${(rate*3600).toFixed(0)}/hr</span>`;}
  else {$('s-eta').textContent='estimating...';}
 }
 $('s-sub').textContent=`· ${DATA.count.toLocaleString()} diagrams across ${Object.keys(DATA.display).length} fields · updated ${fmtTime(DATA.updated_epoch)}`;
}
function render(reset){ if(reset){grid.innerHTML='';shown=0}
 const items=DATA.images.filter(x=>!cat||x.category===cat);
 for(const it of items.slice(shown,shown+PAGE)){
  const d=document.createElement('div');d.className='card';
  d.innerHTML=`<img loading=lazy src="${DATA.base}/${it.url}"><div class=m>${DATA.display[it.category]} · p${it.page}</div>`;
  d.onclick=()=>{$('lbimg').src=`${DATA.base}/${it.url}`;
   const cap=it.caption?`<div style="margin-bottom:8px;line-height:1.5">${it.caption}</div>`:'';
   $('lbc').innerHTML=`${cap}<div style="color:var(--muted);font-size:.85rem">${DATA.display[it.category]} · page ${it.page} · <a href="https://arxiv.org/abs/${it.arxiv_id}" target=_blank>arXiv:${it.arxiv_id}</a> · <a href="${DATA.base}/${it.url}" download>download image</a></div>`;lb.classList.add('on')};
  grid.appendChild(d);}
 shown+=Math.min(PAGE,Math.max(0,items.length-shown));
 more.style.display=shown<items.length?'block':'none';
}
function setCat(c){cat=c;[...cats.children].forEach(b=>b.classList.toggle('on',b.dataset.c===(c||'')));render(true)}
more.onclick=()=>render(false);
function buildCats(){cats.innerHTML='';const counts={};DATA.images.forEach(x=>counts[x.category]=(counts[x.category]||0)+1);
 const mk=(c)=>{const b=document.createElement('button');b.className='cat'+(c===null?' on':'');b.dataset.c=c||'';
  b.textContent=c?`${DATA.display[c]} (${(counts[c]||0).toLocaleString()})`:`All (${DATA.images.length.toLocaleString()})`;
  b.onclick=()=>setCat(c);return b};
 cats.appendChild(mk(null));Object.keys(DATA.display).forEach(c=>cats.appendChild(mk(c)));}
function initTZ(){const sel=$('tz');let zones;
 try{zones=Intl.supportedValuesOf('timeZone')}catch(e){zones=['UTC','America/New_York','America/Chicago','America/Los_Angeles','Europe/London','Europe/Berlin','Asia/Kolkata','Asia/Dubai','Asia/Singapore','Asia/Tokyo','Australia/Sydney']}
 TZ=Intl.DateTimeFormat().resolvedOptions().timeZone;
 if(!zones.includes(TZ))zones.unshift(TZ);
 for(const z of zones){const o=document.createElement('option');o.value=z;o.textContent=z;if(z===TZ)o.selected=true;sel.appendChild(o)}
 sel.onchange=()=>{TZ=sel.value;renderStats();tick()};}
async function refresh(first){
 try{const d=await (await fetch('data.json?t='+Date.now())).json();
  const same=DATA&&DATA.images.length===d.images.length;
  DATA=d;renderStats();
  if(first){buildCats();render(true)}
  else if(!same){const c=cat;buildCats();setCat(c);}
 }catch(e){}
}
initTZ();refresh(true);setInterval(tick,1000);setInterval(()=>refresh(false),60000);
</script></body></html>"""


def _build_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Diagrams"
    cols = [("Field", 26), ("Caption", 90), ("Page", 6), ("arXiv ID", 15),
            ("arXiv URL", 30), ("Image URL", 60), ("Method", 8), ("P(diagram)", 10)]
    for ci, (h, w) in enumerate(cols, 1):
        cell = ws.cell(1, ci, h)
        cell.font = Font(bold=True)
        ws.column_dimensions[chr(64 + ci)].width = w
    wrap = Alignment(wrap_text=True, vertical="top")
    for ri, (name, f, a, p, cap, meth, pd) in enumerate(rows, 2):
        vals = [DISPLAY.get(f, f), cap or "", p, a, f"https://arxiv.org/abs/{a}",
                f"{R2_PUBLIC}/v3/img/{f}/{name}", meth, round(pd or 0, 3)]
        for ci, v in enumerate(vals, 1):
            ws.cell(ri, ci, v).alignment = wrap
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"
    wb.save(path)


def build_and_upload(conn):
    def mget(k, d="0"):
        r = conn.execute("SELECT v FROM meta WHERE k=?", (k,)).fetchone()
        return r[0] if r else d
    rows = conn.execute(
        "SELECT name, field, arxiv_id, page, caption, method, p_diagram FROM images "
        "WHERE status='uploaded' ORDER BY field, name").fetchall()
    count = conn.execute(
        "SELECT COUNT(*) FROM images WHERE status IN ('local','uploaded')").fetchone()[0]
    images = [{"url": f"v3/img/{f}/{n}", "category": f, "arxiv_id": a, "page": p,
               "caption": cap or ""} for (n, f, a, p, cap, _m, _pd) in rows]
    now = time.time()
    data = {"base": R2_PUBLIC, "display": DISPLAY,
            "start": float(mget("pipeline_start", str(now))),
            "target": int(mget("target", "0")),
            "count": count, "updated_epoch": now,
            "images": images}
    gdir = WORK / "gallery"
    gdir.mkdir(parents=True, exist_ok=True)
    (gdir / "data.json").write_text(json.dumps(data))
    (gdir / "index.html").write_text(INDEX_HTML)
    ok1, _ = r2.put(gdir / "data.json", "v3/gallery/data.json", "application/json")
    ok2, _ = r2.put(gdir / "index.html", "v3/gallery/index.html", "text/html")
    # regenerate the export spreadsheet
    try:
        xp = gdir / "stem_diagrams.xlsx"
        _build_xlsx(rows, xp)
        r2.put(xp, "v3/export/stem_diagrams.xlsx",
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as exc:
        import logging
        logging.getLogger("v3").warning("[export] xlsx failed: %s", exc)
    return ok1 and ok2, len(images)
