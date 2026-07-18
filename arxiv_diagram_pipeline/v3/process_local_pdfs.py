#!/usr/bin/env python3
"""Run local PDFs (e.g. textbooks) through the deterministic extractor + gate.

Saves accepted diagrams to a local per-book folder plus one combined Excel.
Private by default — copyrighted material is NOT uploaded anywhere. Reuses
extract.py and gate.py. Run the gate on CPU (GATE_DEVICE=cpu) so it doesn't
contend with the arXiv run on MPS.
"""

import argparse
import csv
import json
import logging
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import extract
import gate

log = logging.getLogger("books")


def process_book(pdf_path, out_dir, threshold, max_pages, chunk=48):
    book = pdf_path.stem
    bdir = out_dir / book
    rows = []
    buf = []  # (page, cand)

    def flush():
        if not buf:
            return
        verdicts = gate.classify([c["image"] for _p, c in buf])
        n = len([r for r in rows])
        for (pno, cd), (cls, p) in zip(buf, verdicts):
            if cls == "diagram" and p >= threshold:
                bdir.mkdir(parents=True, exist_ok=True)
                idx = len(rows) + 1
                name = f"{book[:40].replace(' ', '_')}_p{pno:04d}_{cd['method']}_{idx:04d}.png"
                cd["image"].save(bdir / name)
                rows.append({"book": book, "page": pno, "method": cd["method"],
                             "bbox": json.dumps(cd["bbox"]), "p_diagram": round(p, 3),
                             "file": f"{book}/{name}"})
        buf.clear()

    try:
        for pno, cand in extract.extract_pdf(pdf_path, max_pages=max_pages):
            buf.append((pno, cand))
            if len(buf) >= chunk:
                flush()
        flush()
    except Exception as exc:
        log.warning("  %s: extraction error: %s", book, exc)
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True)
    ap.add_argument("--out", default=str(Path.home() / "Downloads" / "textbook_diagrams"))
    ap.add_argument("--threshold", type=float, default=0.5)
    ap.add_argument("--max-pages", type=int, default=1200)
    ap.add_argument("--limit", type=int, default=0, help="process only first N pdfs (test)")
    args = ap.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")

    src = Path(args.src)
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    pdfs = sorted(src.glob("*.pdf"))
    if args.limit:
        pdfs = pdfs[:args.limit]
    log.info("processing %d PDFs -> %s (gate on %s)", len(pdfs),
             out, __import__("os").getenv("GATE_DEVICE", "mps"))

    all_rows = []
    for i, pdf in enumerate(pdfs, 1):
        t0 = time.time()
        rows = process_book(pdf, out, args.threshold, args.max_pages)
        all_rows.extend(rows)
        log.info("[%d/%d] %.50s -> %d diagrams (%.0fs)", i, len(pdfs),
                 pdf.stem, len(rows), time.time() - t0)
        # incremental manifest so a crash keeps progress
        with open(out / "manifest.csv", "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["book", "page", "method", "bbox", "p_diagram", "file"])
            w.writeheader()
            w.writerows(all_rows)

    # combined Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Textbook Diagrams"
        cols = [("Book", "book", 40), ("Page", "page", 7), ("Method", "method", 10),
                ("P(diagram)", "p_diagram", 11), ("File", "file", 60)]
        for ci, (h, _k, wdt) in enumerate(cols, 1):
            ws.cell(1, ci, h).font = Font(bold=True)
            ws.column_dimensions[chr(64 + ci)].width = wdt
        for ri, r in enumerate(all_rows, 2):
            for ci, (_h, k, _w) in enumerate(cols, 1):
                ws.cell(ri, ci, r[k]).alignment = Alignment(vertical="top")
        ws.freeze_panes = "A2"
        wb.save(out / "textbook_diagrams.xlsx")
    except Exception as exc:
        log.warning("excel failed: %s", exc)

    from collections import Counter
    per = Counter(r["book"] for r in all_rows)
    log.info("DONE: %d diagrams from %d books -> %s", len(all_rows), len(per), out)
    for b, n in per.most_common():
        log.info("   %3d  %s", n, b)


if __name__ == "__main__":
    main()
