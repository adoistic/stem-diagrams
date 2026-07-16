"""Write the two Excel outputs: with-source and without-source."""

import logging
import re

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Control characters Excel refuses (keep \t \n \r); older labels may carry
# them from LaTeX escapes mangled inside LLM JSON output.
_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean(value):
    if isinstance(value, str):
        return _ILLEGAL_XLSX.sub("", value)
    return value

# (header, row key, column width)
WITH_SOURCE_COLUMNS = [
    ("Field", "field", 26),
    ("ArXiv ID", "arxiv_id", 16),
    ("Paper Title", "paper_title", 50),
    ("Authors", "authors", 40),
    ("Published", "published", 12),
    ("Abstract URL", "abs_url", 34),
    ("PDF URL", "pdf_url", 34),
    ("Page", "page_number", 7),
    ("Source Image Path", "source_image_path", 55),
    ("Image File", "image_file", 24),
    ("Diagram Type", "diagram_type", 22),
    ("Diagram Title", "diagram_title", 40),
    ("Label", "label", 110),
]

WITHOUT_SOURCE_COLUMNS = [
    ("Field", "field", 26),
    ("Image File", "image_file", 24),
    ("Diagram Type", "diagram_type", 22),
    ("Diagram Title", "diagram_title", 40),
    ("Label", "label", 110),
]


def _write(rows, columns, dest_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Diagram Labels"
    for col, (header, _key, width) in enumerate(columns, start=1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    wrap = Alignment(wrap_text=True, vertical="top")
    for r, row in enumerate(rows, start=2):
        for c, (_header, key, _w) in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=_clean(row.get(key, "")))
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(dest_path)


def export(rows, output_dir):
    """Write both workbooks; returns their paths."""
    with_path = output_dir / "labels_with_source.xlsx"
    without_path = output_dir / "labels_without_source.xlsx"
    _write(rows, WITH_SOURCE_COLUMNS, with_path)
    _write(rows, WITHOUT_SOURCE_COLUMNS, without_path)
    return with_path, without_path


def _new_write_only_sheet(columns):
    """Build a write_only workbook + sheet with column widths, a bold header
    row, and frozen panes already in place — all of which must be set before
    any data rows are appended in write_only mode."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Diagram Labels")
    for col, (_header, _key, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    try:
        ws.freeze_panes = "A2"
    except (AttributeError, ValueError) as exc:
        log.warning("[export_v2] freeze_panes unsupported, skipping: %s", exc)

    header_cells = []
    for header, _key, _width in columns:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = Font(bold=True)
        header_cells.append(cell)
    ws.append(header_cells)
    return wb, ws


def _append_write_only_row(ws, columns, row, wrap):
    cells = []
    for _header, key, _width in columns:
        cell = WriteOnlyCell(ws, value=_clean(row.get(key, "")))
        cell.alignment = wrap
        cells.append(cell)
    ws.append(cells)


def export_v2(rows_iter, output_dir, total_hint=0):
    """Write labels_with_source.xlsx and labels_without_source.xlsx from an
    iterable of row dicts, using openpyxl write_only workbooks so 30,000+
    rows stream through without holding two workbook copies in memory.

    rows_iter may be a one-pass generator: the with-source workbook is
    written as rows are consumed, while only the 5 without-source column
    values (not the full row dict) are buffered per row so the second
    workbook can be written afterward. Returns (with_path, without_path).
    """
    with_path = output_dir / "labels_with_source.xlsx"
    without_path = output_dir / "labels_without_source.xlsx"
    wrap = Alignment(wrap_text=True, vertical="top")
    without_keys = [key for _header, key, _width in WITHOUT_SOURCE_COLUMNS]

    wb, ws = _new_write_only_sheet(WITH_SOURCE_COLUMNS)
    without_rows = []
    n = 0
    for row in rows_iter:
        _append_write_only_row(ws, WITH_SOURCE_COLUMNS, row, wrap)
        without_rows.append([row.get(key, "") for key in without_keys])
        n += 1
        if total_hint and n % 5000 == 0:
            log.info("[export_v2] with-source: %d/%d rows written", n, total_hint)
    wb.save(with_path)
    log.info("[export_v2] %d rows -> %s", n, with_path)

    wb2, ws2 = _new_write_only_sheet(WITHOUT_SOURCE_COLUMNS)
    for values in without_rows:
        cells = []
        for value in values:
            cell = WriteOnlyCell(ws2, value=_clean(value))
            cell.alignment = wrap
            cells.append(cell)
        ws2.append(cells)
    wb2.save(without_path)
    log.info("[export_v2] %d rows -> %s", len(without_rows), without_path)

    return with_path, without_path
